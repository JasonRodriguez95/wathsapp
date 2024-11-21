from pywhatkit import whats
from openpyxl import load_workbook
import time
import pyautogui as auto

def enviar_mensajes():
    workbook = load_workbook('base.xlsx')
    sheet = workbook.active

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3, values_only=True), start=2):
        try:
            num = str(row[0])
            mensajes = row[1]
            img_path = r"./IMAGEN.PNG"
            wait_time = 20
            
            whats.sendwhats_image(num, img_path, mensajes, wait_time)
            print(f"Mensaje enviado a {num} correctamente.")
            time.sleep(5)
            auto.hotkey('CTRL', 'W')
        except Exception as e:
            print(f"Error al enviar mensaje a {num}: {e}")

    workbook.close()

if __name__ == "__main__":
    enviar_mensajes()